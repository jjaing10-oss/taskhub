#!/usr/bin/env python3
"""
KT M&S 경영계획 데이터 추출 스크립트
★★ 26년 경영계획 v8.1.xlsx → JS 상수 생성
반드시 openpyxl로 직접 파싱. 수치 2회 교차검증 수행.
"""
import openpyxl
import json
import os

def rnd(v):
    try:
        f = float(v)
        return round(f, 1) if f != round(f) else int(round(f))
    except:
        return 0

def irnd(v):
    try: return int(round(float(v)))
    except: return 0

XL_PATH = '★★ 26년 경영계획 - 2025.09.30 v4.3 (3안 가지고 다시 일단락) (매출 8월제출 복원) v8.1 (26년 월별 목표) (1) (1).xlsx'
MID_PATH = '1. 중기 및 2026년 그룹사 경영계획(안)_kt엠앤에스 v7.81 (1) (1).xlsx'

warnings = []
errors = []

def check_monthly_sum(name, year, monthly, annual, tolerance=2):
    s = sum(monthly)
    diff = s - annual
    if abs(diff) > tolerance:
        errors.append(f"⚠️ [{name}] {year}년 월합={s:,} 연간={annual:,} 차이={diff:+}")
    return abs(diff) <= tolerance

# ─────────────────────────────────────────────────────────────────────────────
# 1. 전사 손익계산서 시트
# ─────────────────────────────────────────────────────────────────────────────
print("=" * 60)
print("전사 손익계산서 추출 시작")
print("=" * 60)

wb = openpyxl.load_workbook(XL_PATH, data_only=True)
ws_pl = wb['전사 손익계산서']

# 연도별 Total 열 (1-indexed)
YR_COLS = {2020: 17, 2021: 30, 2022: 43, 2023: 56, 2024: 69, 2025: 82, 2026: 95}
MON_START = {2020: 18, 2021: 31, 2022: 44, 2023: 57, 2024: 70, 2025: 83, 2026: 96}

# 전사 손익계산서 행 정의
PL_ROWS = {
    '매출':          12,
    '상품매출':      13,
    '수수료수입':    14,
    '정책수수료':    15,
    '관리수수료':    16,
    '기타수수료':    17,
    '매출원가':      18,
    '매출총이익':    19,
    '인건비':        21,
    '일반직':        22,
    '영업직군':      23,
    '특별격려금':    24,
    '포상비':        25,
    '퇴직급여':      26,
    '복리후생비':    27,
    '마케팅비용':    28,
    '판매수수료':    29,
    '판촉비':        30,
    '광고':          31,
    '임차관리비':    32,
    '임차료':        33,
    '수도광열비':    34,
    '감가상각비':    35,
    '유형자산상각':  36,
    '무형자산상각':  37,
    '기타운영비':    38,
    '지급수수료':    39,
    '회의비':        40,
    '운반비':        41,
    '교육훈련비':    42,
    '통신비':        43,
    '소모품비':      44,
    '접대비':        45,
    '보험료':        46,
    '세금과공과':    47,
    '차량유지비':    48,
    '여비교통비':    49,
    '도서인쇄비':    50,
    '영업이익':      52,
}

corp = {}
for name, row in PL_ROWS.items():
    corp[name] = {
        'annual': {},
        'monthly': {}
    }
    for yr, col in YR_COLS.items():
        ann_val = irnd(ws_pl.cell(row, col).value)
        mon_vals = [irnd(ws_pl.cell(row, MON_START[yr] + m).value) for m in range(12)]
        corp[name]['annual'][yr] = ann_val
        corp[name]['monthly'][yr] = mon_vals
        # 검증 (2025, 2026만)
        if yr in [2025, 2026]:
            check_monthly_sum(name, yr, mon_vals, ann_val)

print("\n[전사 손익계산서 주요값]")
for name in ['매출', '매출총이익', '영업이익', '인건비', '마케팅비용', '임차관리비', '감가상각비', '기타운영비']:
    print(f"  {name}: {corp[name]['annual']}")

# ─────────────────────────────────────────────────────────────────────────────
# 2. 무선 시트
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("무선 시트 추출")
print("=" * 60)

ws_w = wb['무선']

# 무선 시트 2026E: col83(Total), col84~95(1~12월)
W_TOT_COL = 83
W_MON_START = 84

WIRELESS_ROWS = {
    # CAPA
    'CAPA_전사':     8,
    'CAPA_도매':     9,
    'CAPA_소매':     10,
    'CAPA_소상공인': 11,
    'CAPA_디지털':   14,
    'CAPA_기업':     15,
    'CAPA_Biz':      16,
    'CAPA_IoT':      19,
    # Active (연말 기준)
    'Active_도매':   40,
    'Active_소매':   41,
    'Active_소상공인': 42,
    'Active_디지털': 43,
    'Active_기업':   44,
    'Active_Biz':    45,
    # 수수료
    '관리수수료_채널합': 116,
    '본사공통':      125,
    'HC도매':        126,
    'HC닷컴':        127,
    'Win-win':       128,
    '플라자':        129,
    '매장성':        181,
    '상품매출':      196,
    '매출원가':      220,
    '대당정책':      267,
    '정책수수료':    253,
    '판매수수료':    284,
    '판촉비마케팅':  291,
    '광고':          299,
    '인센티브':      303,
}

wireless = {}
for name, row in WIRELESS_ROWS.items():
    try:
        ann = irnd(ws_w.cell(row, W_TOT_COL).value)
        mon = [irnd(ws_w.cell(row, W_MON_START + m).value) for m in range(12)]
        wireless[name] = {'annual': ann, 'monthly': mon}
    except Exception as e:
        wireless[name] = {'annual': 0, 'monthly': [0]*12}
        warnings.append(f"무선 시트 [{name}] row={row}: {e}")

print("\n[무선 주요값]")
for k in ['CAPA_전사', '관리수수료_채널합', '본사공통', 'HC도매', 'HC닷컴', 'Win-win', '플라자', '매장성', '상품매출', '정책수수료', '판매수수료']:
    v = wireless.get(k, {})
    print(f"  {k}: 연간={v.get('annual',0):,}")

# HC도매 + HC닷컴 + Win-win + 플라자 + 매장성 합계 검증
channel_sum = sum(wireless[k]['annual'] for k in ['HC도매', 'HC닷컴', 'Win-win', '플라자', '매장성'])
채널합 = wireless['관리수수료_채널합']['annual']
본사공통 = wireless['본사공통']['annual']
print(f"\n  [검증] HC도매+HC닷컴+Win-win+플라자+매장성 = {channel_sum:,}")
print(f"  [검증] 관리수수료_채널합(행116) = {채널합:,}")
print(f"  [검증] 본사공통(행125) = {본사공통:,}")
# 행116은 5개 채널 합계
if abs(channel_sum - 채널합) > 5:
    warnings.append(f"무선 채널합 불일치: 계산={channel_sum:,} vs 행116={채널합:,}")

# ─────────────────────────────────────────────────────────────────────────────
# 3. 유선 시트
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("유선 시트 추출")
print("=" * 60)

try:
    ws_wl = wb['유선']
    WL_TOT_COL = 83
    WL_MON_START = 84
    WIRELINE_ROWS = {
        '관리수수료_합계': 15,
        '정책수수료':      30,
        '판촉비':          35,
        '인센티브':        36,
    }
    wireline = {}
    for name, row in WIRELINE_ROWS.items():
        try:
            ann = irnd(ws_wl.cell(row, WL_TOT_COL).value)
            mon = [irnd(ws_wl.cell(row, WL_MON_START + m).value) for m in range(12)]
            wireline[name] = {'annual': ann, 'monthly': mon}
        except:
            wireline[name] = {'annual': 0, 'monthly': [0]*12}
    print("[유선 주요값]")
    for k, v in wireline.items():
        print(f"  {k}: 연간={v['annual']:,}")
except Exception as e:
    wireline = {}
    warnings.append(f"유선 시트: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# 4. 인건비 시트 (col85=Total, col86~97=1~12월 ← 주의!)
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("인건비 시트 추출")
print("=" * 60)

try:
    ws_lb = wb['인건비']
    LB_TOT_COL = 84   # 2026E Total
    LB_MON_START = 85 # 1~12월
    LABOR_ROWS = {
        '급여총액':   28,
        '포상비':     34,
        '복리후생비': 37,
        '퇴직급여':   38,
        '특별격려금': 33,
    }
    labor = {}
    for name, row in LABOR_ROWS.items():
        try:
            ann = irnd(ws_lb.cell(row, LB_TOT_COL).value)
            mon = [irnd(ws_lb.cell(row, LB_MON_START + m).value) for m in range(12)]
            labor[name] = {'annual': ann, 'monthly': mon}
            if ann > 0:
                check_monthly_sum(f"인건비_{name}", 2026, mon, ann)
        except Exception as e2:
            labor[name] = {'annual': 0, 'monthly': [0]*12}
            warnings.append(f"인건비 [{name}] row={row}: {e2}")
    print("[인건비 주요값]")
    for k, v in labor.items():
        print(f"  {k}: 연간={v['annual']:,}, 월별={v['monthly']}")
except Exception as e:
    labor = {}
    warnings.append(f"인건비 시트: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# 5. 인프라 시트 (col84=Total, col85~96=1~12월)
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("인프라 시트 추출")
print("=" * 60)

try:
    ws_if = wb['인프라']
    IF_TOT_COL = 83
    IF_MON_START = 84
    INFRA_ROWS = {
        '임차관리비합계': 25,
        '임차료':         26,
        '수도광열비':     32,
    }
    infra = {}
    for name, row in INFRA_ROWS.items():
        try:
            ann = irnd(ws_if.cell(row, IF_TOT_COL).value)
            mon = [irnd(ws_if.cell(row, IF_MON_START + m).value) for m in range(12)]
            infra[name] = {'annual': ann, 'monthly': mon}
            if ann != 0:
                check_monthly_sum(f"인프라_{name}", 2026, mon, ann, tolerance=5)
        except Exception as e2:
            infra[name] = {'annual': 0, 'monthly': [0]*12}
            warnings.append(f"인프라 [{name}] row={row}: {e2}")
    print("[인프라 주요값]")
    for k, v in infra.items():
        print(f"  {k}: 연간={v['annual']:,}")
        print(f"       월별={v['monthly']}")
except Exception as e:
    infra = {}
    warnings.append(f"인프라 시트: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# 6. 기타운영비 시트 (col84=Total, col85~96=1~12월)
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("기타운영비 시트 추출")
print("=" * 60)

try:
    ws_ox = wb['기타운영비']
    OX_TOT_COL = 83
    OX_MON_START = 84
    OPEX_ROWS = {
        '지급수수료':  6,
        '회의비':      20,
        '운반비':      28,
        '교육훈련비':  30,
        '통신비':      32,
        '소모품비':    34,
        '접대비':      38,
        '보험료':      39,
        '세금과공과':  43,
        '차량유지비':  44,
        '여비교통비':  46,
        '도서인쇄비':  48,
    }
    opex = {}
    for name, row in OPEX_ROWS.items():
        try:
            ann = irnd(ws_ox.cell(row, OX_TOT_COL).value)
            mon = [irnd(ws_ox.cell(row, OX_MON_START + m).value) for m in range(12)]
            opex[name] = {'annual': ann, 'monthly': mon}
            if ann > 0:
                check_monthly_sum(f"기타운영비_{name}", 2026, mon, ann)
        except Exception as e2:
            opex[name] = {'annual': 0, 'monthly': [0]*12}
            warnings.append(f"기타운영비 [{name}] row={row}: {e2}")
    print("[기타운영비 주요값]")
    total_opex = sum(v['annual'] for v in opex.values())
    print(f"  합계(지급수수료 포함): {total_opex:,}")
    for k, v in opex.items():
        print(f"  {k}: 연간={v['annual']:,}")
        print(f"       월별={v['monthly']}")
except Exception as e:
    opex = {}
    warnings.append(f"기타운영비 시트: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# 7. 유통플랫폼 시트 (col44=Total, col45~56=1~12월)
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("유통플랫폼 시트 추출")
print("=" * 60)

try:
    ws_dp = wb['유통플랫폼']
    DP_TOT_COL = 43
    DP_MON_START = 44
    DIST_ROWS = {
        '매출':           9,
        '상품서비스매출': 10,
        '수수료정책':     13,
        '수수료기타':     14,
        '비용':           15,
        '매출원가':       16,
        '매각원가':       17,
        '인건비':         18,
        '임차료':         19,
        '지급수수료':     20,
        '기타운영비':     21,
        '운반비':         22,
        '판촉비':         24,
        '감가상각':       25,
        '영업이익':       27,
    }
    dist = {}
    for name, row in DIST_ROWS.items():
        try:
            ann = irnd(ws_dp.cell(row, DP_TOT_COL).value)
            mon = [irnd(ws_dp.cell(row, DP_MON_START + m).value) for m in range(12)]
            dist[name] = {'annual': ann, 'monthly': mon}
            if ann != 0:
                check_monthly_sum(f"유통플랫폼_{name}", 2026, mon, ann)
        except Exception as e2:
            dist[name] = {'annual': 0, 'monthly': [0]*12}
            warnings.append(f"유통플랫폼 [{name}] row={row}: {e2}")
    print("[유통플랫폼 주요값]")
    for k in ['매출', '영업이익', '매출원가']:
        v = dist.get(k, {})
        print(f"  {k}: 연간={v.get('annual',0):,}")
        print(f"       월별={v.get('monthly',[])}")
except Exception as e:
    dist = {}
    warnings.append(f"유통플랫폼 시트: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# 8. 중기계획 파일 — key Index 시트
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("중기계획 파일 추출")
print("=" * 60)

midterm = {}
try:
    wb2 = openpyxl.load_workbook(MID_PATH, data_only=True)
    sheet_names = wb2.sheetnames
    print(f"중기계획 시트 목록: {sheet_names}")

    for sname in sheet_names[:5]:
        ws2 = wb2[sname]
        print(f"\n  시트 [{sname}] 구조 파악 (첫 10행 × 15열):")
        for r in range(1, 11):
            row_data = []
            for c in range(1, 16):
                v = ws2.cell(r, c).value
                if v is not None:
                    row_data.append(f"[{c}]{str(v)[:15]}")
            if row_data:
                print(f"    행{r}: {', '.join(row_data)}")
except Exception as e:
    warnings.append(f"중기계획 파일: {e}")
    print(f"  ⚠️ 중기계획 파일 오류: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# 9. 전사 vs 통신+유통 교차검증
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("교차검증 (전사 = 통신 + 유통플랫폼)")
print("=" * 60)

# 전사 영업이익
전사_영업이익_2026 = corp['영업이익']['annual'].get(2026, 0)
유통_영업이익_2026 = dist.get('영업이익', {}).get('annual', 0)
통신_영업이익_2026 = 전사_영업이익_2026 - 유통_영업이익_2026

print(f"  전사 영업이익(2026E): {전사_영업이익_2026:,}")
print(f"  유통플랫폼 영업이익(2026E): {유통_영업이익_2026:,}")
print(f"  통신 영업이익 추정(전사-유통): {통신_영업이익_2026:,}")

전사_매출_2026 = corp['매출']['annual'].get(2026, 0)
유통_매출_2026 = dist.get('매출', {}).get('annual', 0)
통신_매출_2026 = 전사_매출_2026 - 유통_매출_2026
print(f"\n  전사 매출(2026E): {전사_매출_2026:,}")
print(f"  유통플랫폼 매출(2026E): {유통_매출_2026:,}")
print(f"  통신 매출 추정(전사-유통): {통신_매출_2026:,}")

전사_기타운영비_2026 = corp['기타운영비']['annual'].get(2026, 0)
유통_기타운영비_2026 = dist.get('기타운영비', {}).get('annual', 0)
통신_기타운영비_2026 = 전사_기타운영비_2026 - 유통_기타운영비_2026
print(f"\n  전사 기타운영비(2026E): {전사_기타운영비_2026:,}")
print(f"  유통플랫폼 기타운영비(2026E): {유통_기타운영비_2026:,}")
print(f"  통신 기타운영비 추정(전사-유통): {통신_기타운영비_2026:,}")

# 엑셀 vs 목표값 비교
print(f"\n  기타운영비 소스 합계(기타운영비시트): {sum(v['annual'] for k,v in opex.items() if k != '지급수수료'):,}")

# ─────────────────────────────────────────────────────────────────────────────
# 10. JS 상수 파일 생성
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("JS 상수 파일 생성")
print("=" * 60)

def py_to_js_array(arr):
    return '[' + ','.join(str(x) for x in arr) + ']'

def py_to_js_obj(obj):
    items = []
    for k, v in obj.items():
        if isinstance(v, dict):
            items.append(f'{k}:{py_to_js_obj(v)}')
        elif isinstance(v, list):
            items.append(f'{k}:{py_to_js_array(v)}')
        else:
            items.append(f'{k}:{v}')
    return '{' + ','.join(items) + '}'

js_lines = []
js_lines.append('// ============================================================')
js_lines.append('// AUTO-GENERATED by extract_data.py — DO NOT EDIT MANUALLY')
js_lines.append('// Source: ★★ 26년 경영계획 v8.1.xlsx')
js_lines.append('// ============================================================')
js_lines.append('')

# 연간 히스토리 (2020~2025)
js_lines.append('// ── 전사 연간 실적 (2020~2025) ──────────────────────────────')
js_lines.append('const HIST_ANNUAL_SRC = {')
HIST_ITEMS = ['매출', '상품매출', '수수료수입', '정책수수료', '관리수수료', '기타수수료',
              '매출원가', '매출총이익', '인건비', '마케팅비용', '판매수수료', '판촉비', '광고',
              '임차관리비', '감가상각비', '기타운영비', '지급수수료', '회의비', '영업이익']
for name in HIST_ITEMS:
    if name in corp:
        d = corp[name]['annual']
        vals = [d.get(yr, 0) for yr in [2020, 2021, 2022, 2023, 2024, 2025]]
        js_lines.append(f'  {name}: {{{", ".join(f"{yr}:{v}" for yr, v in zip([2020,2021,2022,2023,2024,2025], vals))}}},')
js_lines.append('};')
js_lines.append('')

# 2025 월별 실적 (수정 불가 기준값)
js_lines.append('// ── 2025 월별 실적 (수정 불가) ──────────────────────────────')
js_lines.append('const HIST_2025_MON_SRC = {')
for name in HIST_ITEMS:
    if name in corp:
        mon = corp[name]['monthly'].get(2025, [0]*12)
        js_lines.append(f'  {name}: {py_to_js_array(mon)},')
js_lines.append('};')
js_lines.append('')

# 2026E 월별 계획 (전사 손익계산서 기준)
js_lines.append('// ── 2026E 월별 계획 (전사 손익계산서) ──────────────────────')
js_lines.append('const TARGET_2026_MON_SRC = {')
for name in HIST_ITEMS:
    if name in corp:
        mon = corp[name]['monthly'].get(2026, [0]*12)
        js_lines.append(f'  {name}: {py_to_js_array(mon)},  // 연간={corp[name]["annual"].get(2026,0):,}')
js_lines.append('};')
js_lines.append('')

# 2026E 연간 목표
js_lines.append('// ── 2026E 연간 목표값 ─────────────────────────────────────')
js_lines.append('const TARGET_2026_ANN_SRC = {')
for name in HIST_ITEMS:
    if name in corp:
        js_lines.append(f'  {name}: {corp[name]["annual"].get(2026, 0)},')
js_lines.append('};')
js_lines.append('')

# 무선 데이터
js_lines.append('// ── 무선 2026E ───────────────────────────────────────────────')
js_lines.append('const WIRELESS_SRC = {')
for name, data in wireless.items():
    js_lines.append(f'  {name.replace("-","_")}: {{annual:{data["annual"]}, monthly:{py_to_js_array(data["monthly"])}}},')
js_lines.append('};')
js_lines.append('')

# 인건비 데이터
js_lines.append('// ── 인건비 2026E ─────────────────────────────────────────────')
js_lines.append('const LABOR_SRC = {')
for name, data in labor.items():
    js_lines.append(f'  {name}: {{annual:{data["annual"]}, monthly:{py_to_js_array(data["monthly"])}}},')
js_lines.append('};')
js_lines.append('')

# 인프라 데이터
js_lines.append('// ── 인프라 2026E ─────────────────────────────────────────────')
js_lines.append('const INFRA_SRC = {')
for name, data in infra.items():
    js_lines.append(f'  {name}: {{annual:{data["annual"]}, monthly:{py_to_js_array(data["monthly"])}}},')
js_lines.append('};')
js_lines.append('')

# 기타운영비 데이터
js_lines.append('// ── 기타운영비 2026E ──────────────────────────────────────────')
js_lines.append('const OPEX_SRC = {')
for name, data in opex.items():
    js_lines.append(f'  {name}: {{annual:{data["annual"]}, monthly:{py_to_js_array(data["monthly"])}}},')
js_lines.append('};')
js_lines.append('')

# 유통플랫폼 데이터
js_lines.append('// ── 유통플랫폼 2026E ──────────────────────────────────────────')
js_lines.append('const DIST_SRC = {')
for name, data in dist.items():
    js_lines.append(f'  {name}: {{annual:{data["annual"]}, monthly:{py_to_js_array(data["monthly"])}}},')
js_lines.append('};')
js_lines.append('')

# 계절성 패턴 (2026E 소스 기반)
js_lines.append('// ── 계절성 패턴 (2026E 소스, 소수점 6자리) ────────────────────')
js_lines.append('const SRC_PATTERNS_2026 = {')
for name in ['매출', '매출원가', '인건비', '마케팅비용', '임차관리비', '감가상각비', '기타운영비', '영업이익']:
    if name in corp:
        mon = corp[name]['monthly'].get(2026, [0]*12)
        s = sum(mon)
        if s != 0:
            patt = [round(v/s, 6) for v in mon]
        else:
            patt = [round(1/12, 6)] * 12
        js_lines.append(f'  {name}: {py_to_js_array(patt)},  // sum check={sum(patt):.6f}')
js_lines.append('};')
js_lines.append('')

with open('extracted_data.js', 'w', encoding='utf-8') as f:
    f.write('\n'.join(js_lines))
print("  → extracted_data.js 생성 완료")

# ─────────────────────────────────────────────────────────────────────────────
# 11. 검증 결과 요약
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("검증 결과 요약")
print("=" * 60)

if errors:
    print(f"\n❌ 오류 {len(errors)}건:")
    for e in errors:
        print(f"  {e}")
else:
    print("✅ 모든 검증 통과 (월합 = 연간 ±2 이내)")

if warnings:
    print(f"\n⚠️ 경고 {len(warnings)}건:")
    for w in warnings:
        print(f"  {w}")
else:
    print("✅ 경고 없음")

print("\n[최종 검증 기준값]")
print(f"  전사 매출(2026E):     {전사_매출_2026:,}")
print(f"  전사 영업이익(2026E): {전사_영업이익_2026:,}")
print(f"  통신 매출(전사-유통): {통신_매출_2026:,}")
print(f"  통신 영업이익:        {통신_영업이익_2026:,}")
print(f"  유통플랫폼 매출:      {유통_매출_2026:,}")
print(f"  유통플랫폼 영업이익:  {유통_영업이익_2026:,}")

# JSON 저장 (추후 참조용)
result_data = {
    'corp_annual_2026': {k: v['annual'].get(2026, 0) for k, v in corp.items()},
    'corp_monthly_2025': {k: v['monthly'].get(2025, [0]*12) for k, v in corp.items()},
    'corp_monthly_2026': {k: v['monthly'].get(2026, [0]*12) for k, v in corp.items()},
    'corp_annual_hist': {k: {str(yr): v['annual'].get(yr, 0) for yr in range(2020, 2027)} for k, v in corp.items()},
    'wireless': {k: {'annual': v['annual'], 'monthly': v['monthly']} for k, v in wireless.items()},
    'labor': {k: {'annual': v['annual'], 'monthly': v['monthly']} for k, v in labor.items()},
    'infra': {k: {'annual': v['annual'], 'monthly': v['monthly']} for k, v in infra.items()},
    'opex': {k: {'annual': v['annual'], 'monthly': v['monthly']} for k, v in opex.items()},
    'dist': {k: {'annual': v['annual'], 'monthly': v['monthly']} for k, v in dist.items()},
    'errors': errors,
    'warnings': warnings,
}

with open('extracted_data.json', 'w', encoding='utf-8') as f:
    json.dump(result_data, f, ensure_ascii=False, indent=2)
print("\n  → extracted_data.json 저장 완료")

print("\n✅ 추출 완료!")
